"""Model-free artifact profiling for common research file formats."""

from __future__ import annotations

import csv
import json
import re
import sqlite3
import zipfile
from dataclasses import asdict, dataclass
from pathlib import Path
from typing import Any


@dataclass(frozen=True)
class ArtifactProfile:
    path: str
    kind: str
    size_bytes: int
    shape: dict[str, Any]
    schema: list[dict[str, Any]]
    missingness: dict[str, int]
    formulas: list[str]
    references: list[str]
    preview: Any
    warnings: list[str]

    def to_dict(self) -> dict[str, Any]:
        return asdict(self)


def _infer(values: list[Any]) -> str:
    present = [value for value in values if value not in (None, "")]
    if not present:
        return "unknown"
    if all(isinstance(value, bool) for value in present):
        return "boolean"
    if all(isinstance(value, (int, float)) and not isinstance(value, bool) for value in present):
        return "number"
    return "string"


def _table_profile(path: Path, rows: list[dict[str, Any]], kind: str, total: int | None = None) -> ArtifactProfile:
    columns = sorted({str(key) for row in rows for key in row})
    schema = []
    missingness: dict[str, int] = {}
    for column in columns:
        values = [row.get(column) for row in rows]
        missingness[column] = sum(value in (None, "") for value in values)
        schema.append({"name": column, "type": _infer(values), "nullable": missingness[column] > 0})
    return ArtifactProfile(
        path=str(path), kind=kind, size_bytes=path.stat().st_size,
        shape={"rows": total if total is not None else len(rows), "columns": len(columns)},
        schema=schema, missingness=missingness, formulas=[], references=[],
        preview=rows[:5], warnings=[] if rows else ["No data rows found."],
    )


def _delimited(path: Path) -> ArtifactProfile:
    with path.open(newline="", encoding="utf-8-sig", errors="replace") as handle:
        reader = csv.DictReader(handle, delimiter="\t" if path.suffix.lower() == ".tsv" else ",")
        # strict=False is intentional: this caps the sample at 1000 rows,
        # and a shorter file legitimately exhausts the reader first.
        rows = [dict(row) for _, row in zip(range(1000), reader, strict=False)]
    return _table_profile(path, rows, "tsv" if path.suffix.lower() == ".tsv" else "csv")


def _json_profile(path: Path) -> ArtifactProfile:
    payload = json.loads(path.read_text(encoding="utf-8"))
    if isinstance(payload, list) and all(isinstance(row, dict) for row in payload):
        return _table_profile(path, payload[:1000], "json_records", len(payload))
    references = sorted(set(re.findall(r"(?:doi|arxiv|https?://)[^\s\"']+", path.read_text(encoding="utf-8"))))
    return ArtifactProfile(
        path=str(path), kind="json", size_bytes=path.stat().st_size,
        shape={"type": type(payload).__name__}, schema=[], missingness={},
        formulas=[], references=references[:50], preview=payload, warnings=[],
    )


def _sqlite_profile(path: Path) -> ArtifactProfile:
    tables: list[dict[str, Any]] = []
    with sqlite3.connect(path) as connection:
        names = [row[0] for row in connection.execute("SELECT name FROM sqlite_master WHERE type='table' ORDER BY name")]
        for name in names:
            columns = [
                {"cid": row[0], "name": row[1], "type": row[2], "nullable": not bool(row[3]), "primary_key": bool(row[5])}
                for row in connection.execute(f'PRAGMA table_info("{name}")').fetchall()
            ]
            count = connection.execute(f'SELECT COUNT(*) FROM "{name}"').fetchone()[0]
            tables.append({"name": name, "rows": count, "columns": columns})
    return ArtifactProfile(
        path=str(path), kind="sqlite", size_bytes=path.stat().st_size,
        shape={"tables": len(tables)}, schema=tables, missingness={}, formulas=[],
        references=[], preview=tables[:5], warnings=[],
    )


def _parquet_profile(path: Path) -> ArtifactProfile:
    try:
        import pyarrow.parquet as parquet

        table = parquet.read_table(path, use_threads=False)
        schema = [{"name": field.name, "type": str(field.type), "nullable": field.nullable} for field in table.schema]
        preview = table.slice(0, 5).to_pylist()
        missingness = {field["name"]: sum(row.get(field["name"]) is None for row in preview) for field in schema}
        return ArtifactProfile(
            path=str(path), kind="parquet", size_bytes=path.stat().st_size,
            shape={"rows": table.num_rows, "columns": table.num_columns}, schema=schema,
            missingness=missingness, formulas=[], references=[], preview=preview, warnings=[],
        )
    except Exception as exc:  # noqa: BLE001 - profile remains useful without pyarrow
        return ArtifactProfile(
            path=str(path), kind="parquet", size_bytes=path.stat().st_size,
            shape={}, schema=[], missingness={}, formulas=[], references=[], preview=[],
            warnings=[f"Parquet reader unavailable: {exc}"],
        )


def _text_profile(path: Path, kind: str) -> ArtifactProfile:
    text = path.read_text(encoding="utf-8", errors="replace")
    formulas = sorted(set(re.findall(r"(?m)^\s*\\(?:begin|end)\{[^}]+\}|\$[^$]+\$", text)))[:50]
    references = sorted(set(re.findall(r"(?:doi:\s*|https?://|\\cite\{)[^\s}]+", text, flags=re.IGNORECASE)))[:50]
    return ArtifactProfile(
        path=str(path), kind=kind, size_bytes=path.stat().st_size,
        shape={"lines": len(text.splitlines()), "characters": len(text)}, schema=[],
        missingness={}, formulas=formulas, references=references, preview=text[:2000], warnings=[],
    )


def _image_profile(path: Path) -> ArtifactProfile:
    shape: dict[str, Any] = {}
    warnings: list[str] = []
    try:
        from PIL import Image

        with Image.open(path) as image:
            shape = {"width": image.width, "height": image.height, "mode": image.mode}
    except Exception as exc:  # noqa: BLE001
        warnings.append(f"Image metadata reader unavailable: {exc}")
    return ArtifactProfile(
        path=str(path), kind="image", size_bytes=path.stat().st_size,
        shape=shape, schema=[], missingness={}, formulas=[], references=[], preview=None, warnings=warnings,
    )


def _spreadsheet_profile(path: Path) -> ArtifactProfile:
    warnings: list[str] = []
    sheets: list[dict[str, Any]] = []
    formulas: list[str] = []
    try:
        import openpyxl

        workbook = openpyxl.load_workbook(path, read_only=True, data_only=False)
        for sheet in workbook.worksheets:
            rows = list(sheet.iter_rows(values_only=False, max_row=1000))
            values = [[cell.value for cell in row] for row in rows]
            formulas.extend(str(value) for row in values for value in row if isinstance(value, str) and value.startswith("="))
            sheets.append({"name": sheet.title, "rows": sheet.max_row, "columns": sheet.max_column, "preview": values[:5]})
    except Exception as exc:  # noqa: BLE001
        warnings.append(f"Spreadsheet reader unavailable: {exc}")
    return ArtifactProfile(
        path=str(path), kind="spreadsheet", size_bytes=path.stat().st_size,
        shape={"sheets": len(sheets)}, schema=sheets, missingness={}, formulas=formulas[:100],
        references=[], preview=sheets[:5], warnings=warnings,
    )


def profile_path(path: str | Path) -> ArtifactProfile:
    """Profile an artifact without executing or sending its contents anywhere."""
    resolved = Path(path).expanduser().resolve()
    if not resolved.is_file():
        raise ValueError(f"Artifact path is not a file: {path}")
    suffix = resolved.suffix.lower()
    if suffix in {".csv", ".tsv"}:
        return _delimited(resolved)
    if suffix == ".json":
        return _json_profile(resolved)
    if suffix in {".sqlite", ".sqlite3", ".db"}:
        return _sqlite_profile(resolved)
    if suffix == ".parquet":
        return _parquet_profile(resolved)
    if suffix in {".png", ".jpg", ".jpeg", ".tif", ".tiff", ".webp", ".bmp"}:
        return _image_profile(resolved)
    if suffix in {".xlsx", ".xlsm"}:
        return _spreadsheet_profile(resolved)
    if suffix in {".tex", ".md", ".txt", ".ipynb", ".py", ".js", ".rs"}:
        return _text_profile(resolved, "notebook" if suffix == ".ipynb" else "text")
    if suffix == ".pdf":
        return ArtifactProfile(str(resolved), "pdf", resolved.stat().st_size, {}, [], {}, [], [], None, ["Use paper-visual for page-level PDF evidence."])
    if zipfile.is_zipfile(resolved):
        return _text_profile(resolved, "archive")
    return _text_profile(resolved, "text")
